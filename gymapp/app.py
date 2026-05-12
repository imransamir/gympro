from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
from datetime import datetime, timedelta
import csv
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gym.db')

# ─── DATABASE ─────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS trainers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            phone TEXT,
            specializations TEXT,
            hire_date TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            duration INTEGER,
            max_capacity INTEGER,
            category TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trainer_id INTEGER NOT NULL,
            class_id INTEGER NOT NULL,
            assigned_date DATE NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            status TEXT DEFAULT 'scheduled',
            FOREIGN KEY (trainer_id) REFERENCES trainers(id),
            FOREIGN KEY (class_id) REFERENCES classes(id)
        );
        CREATE TABLE IF NOT EXISTS working_hours (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trainer_id INTEGER NOT NULL,
            assignment_id INTEGER,
            date DATE NOT NULL,
            hours_worked REAL NOT NULL,
            description TEXT,
            recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (trainer_id) REFERENCES trainers(id),
            FOREIGN KEY (assignment_id) REFERENCES assignments(id)
        );
    ''')
    conn.commit()

    # Migrate: add columns that may be missing from older schema
    existing = {row[1] for row in c.execute("PRAGMA table_info(trainers)")}
    if 'status' not in existing:
        c.execute("ALTER TABLE trainers ADD COLUMN status TEXT DEFAULT 'active'")
        conn.commit()

    # Seed data only if empty
    c.execute("SELECT COUNT(*) FROM trainers")
    if c.fetchone()[0] == 0:
        trainers = [
            ('James Carter',   'james@gympro.com',   '07700 900001', 'Yoga, Mindfulness',         '2022-03-15'),
            ('Sarah Mitchell', 'sarah@gympro.com',   '07700 900002', 'Weight Training, Cardio',   '2021-06-01'),
            ('Mike Thompson',  'mike@gympro.com',    '07700 900003', 'CrossFit, HIIT',             '2023-01-10'),
            ('Emma Wilson',    'emma@gympro.com',    '07700 900004', 'Pilates, Stretching',        '2022-09-20'),
            ('David Brown',    'david@gympro.com',   '07700 900005', 'Boxing, Kickboxing',         '2020-11-05'),
            ('Lisa Johnson',   'lisa@gympro.com',    '07700 900006', 'Zumba, Dance Fitness',       '2023-04-22'),
            ('Ryan Adams',     'ryan@gympro.com',    '07700 900007', 'Powerlifting, Strength',     '2021-08-14'),
            ('Sophie Turner',  'sophie@gympro.com',  '07700 900008', 'Spin, Cycling',              '2022-12-01'),
            ('Chris Evans',    'chris@gympro.com',   '07700 900009', 'Swimming, Aqua Aerobics',    '2020-05-18'),
            ('Nina Patel',     'nina@gympro.com',    '07700 900010', 'Nutrition, Body Conditioning','2023-07-30'),
        ]
        c.executemany(
            "INSERT INTO trainers (name,email,phone,specializations,hire_date) VALUES(?,?,?,?,?)",
            trainers
        )
        classes = [
            ('Morning Yoga',       'Relaxing morning flow',       60, 20, 'Mind & Body'),
            ('Weight Training',    'Strength building session',   90, 15, 'Strength'),
            ('HIIT Cardio',        'High intensity intervals',    45, 25, 'Cardio'),
            ('Pilates Core',       'Core strength & flexibility', 55, 18, 'Mind & Body'),
            ('Boxing Fundamentals','Technique & conditioning',    60, 12, 'Combat'),
            ('Zumba Party',        'Dance-based cardio',          50, 30, 'Dance'),
            ('Spin Cycle',         'Indoor cycling class',        45, 20, 'Cardio'),
            ('Powerlifting',       'Barbell strength training',   75, 10, 'Strength'),
            ('Aqua Aerobics',      'Low-impact pool workout',     50, 15, 'Aqua'),
            ('Body Conditioning',  'Full body toning',            60, 22, 'Fitness'),
        ]
        c.executemany(
            "INSERT INTO classes (name,description,duration,max_capacity,category) VALUES(?,?,?,?,?)",
            classes
        )
        conn.commit()
    conn.close()

# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

# --- Trainers ---
@app.route('/api/trainers', methods=['GET'])
def get_trainers():
    conn = get_db()
    rows = conn.execute("SELECT * FROM trainers ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/trainers', methods=['POST'])
def create_trainer():
    d = request.json
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO trainers (name,email,phone,specializations,hire_date,status) VALUES(?,?,?,?,?,?)",
            (d['name'], d['email'], d.get('phone',''), d.get('specializations',''), d.get('hire_date', datetime.now().strftime('%Y-%m-%d')), 'active')
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/trainers/<int:tid>', methods=['DELETE'])
def delete_trainer(tid):
    conn = get_db()
    conn.execute("DELETE FROM trainers WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# --- Classes ---
@app.route('/api/classes', methods=['GET'])
def get_classes():
    conn = get_db()
    rows = conn.execute("SELECT * FROM classes ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# --- Assignments ---
@app.route('/api/assignments', methods=['GET'])
def get_assignments():
    conn = get_db()
    rows = conn.execute('''
        SELECT a.id, t.name as trainer, c.name as class_name, c.category,
               a.assigned_date, a.start_time, a.end_time, a.status
        FROM assignments a
        JOIN trainers t ON a.trainer_id=t.id
        JOIN classes c ON a.class_id=c.id
        ORDER BY a.assigned_date DESC, a.start_time
        LIMIT 50
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/assignments', methods=['POST'])
def create_assignment():
    d = request.json
    if not d:
        return jsonify({'success': False, 'error': 'Invalid request body'}), 400
    try:
        trainer_id = int(d['trainer_id'])
        class_id   = int(d['class_id'])
        date       = d['date']
        start      = d['start_time']
        end        = d['end_time']
    except (KeyError, ValueError, TypeError) as e:
        return jsonify({'success': False, 'error': f'Missing or invalid field: {e}'}), 400

    if not date or not start or not end:
        return jsonify({'success': False, 'error': 'Date and times are required'}), 400

    conn = get_db()
    # Standard interval overlap: A overlaps B iff A.start < B.end AND A.end > B.start
    conflict = conn.execute('''
        SELECT COUNT(*) FROM assignments
        WHERE trainer_id=? AND assigned_date=?
        AND start_time < ? AND end_time > ?
    ''', (trainer_id, date, end, start)).fetchone()[0]

    if conflict:
        conn.close()
        return jsonify({'success': False, 'error': 'Trainer has a scheduling conflict at this time.'}), 409

    conn.execute(
        "INSERT INTO assignments (trainer_id,class_id,assigned_date,start_time,end_time) VALUES(?,?,?,?,?)",
        (trainer_id, class_id, date, start, end)
    )
    assignment_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Auto-record hours (SCRUM-17)
    try:
        fmt = "%H:%M"
        hrs = (datetime.strptime(end, fmt) - datetime.strptime(start, fmt)).seconds / 3600
    except Exception:
        hrs = 1.0

    conn.execute(
        "INSERT INTO working_hours (trainer_id,assignment_id,date,hours_worked,description) VALUES(?,?,?,?,?)",
        (trainer_id, assignment_id, date, hrs, "Auto-recorded for class assignment")
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# --- Working Hours ---
@app.route('/api/hours', methods=['GET'])
def get_hours():
    conn = get_db()
    rows = conn.execute('''
        SELECT wh.id, t.name as trainer, wh.date, wh.hours_worked, wh.description,
               c.name as class_name, wh.recorded_at
        FROM working_hours wh
        JOIN trainers t ON wh.trainer_id=t.id
        LEFT JOIN assignments a ON wh.assignment_id=a.id
        LEFT JOIN classes c ON a.class_id=c.id
        ORDER BY wh.date DESC
        LIMIT 100
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/hours/manual', methods=['POST'])
def record_hours_manual():
    d = request.json
    conn = get_db()
    conn.execute(
        "INSERT INTO working_hours (trainer_id,date,hours_worked,description) VALUES(?,?,?,?)",
        (d['trainer_id'], d['date'], d['hours'], d.get('description', 'Manual entry'))
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# --- Reports ---
@app.route('/api/report', methods=['GET'])
def get_report():
    trainer_id = request.args.get('trainer_id')
    start_date = request.args.get('start_date')
    end_date   = request.args.get('end_date')

    query = '''
        SELECT t.name, t.specializations,
               ROUND(SUM(wh.hours_worked), 2) as total_hours,
               COUNT(DISTINCT a.id) as classes_taught,
               ROUND(AVG(wh.hours_worked), 2) as avg_hours
        FROM trainers t
        JOIN working_hours wh ON t.id=wh.trainer_id
        LEFT JOIN assignments a ON wh.assignment_id=a.id
    '''
    conds, params = [], []
    if trainer_id:
        conds.append("t.id=?"); params.append(trainer_id)
    if start_date:
        conds.append("wh.date>=?"); params.append(start_date)
    if end_date:
        conds.append("wh.date<=?"); params.append(end_date)
    if conds:
        query += " WHERE " + " AND ".join(conds)
    query += " GROUP BY t.id, t.name ORDER BY total_hours DESC"

    conn = get_db()
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/report/export')
def export_report():
    fmt        = request.args.get('format', 'csv')
    trainer_id = request.args.get('trainer_id')
    start_date = request.args.get('start_date')
    end_date   = request.args.get('end_date')

    query = '''
        SELECT t.name, t.specializations,
               ROUND(SUM(wh.hours_worked),2) as total_hours,
               COUNT(DISTINCT a.id) as classes_taught,
               ROUND(AVG(wh.hours_worked),2) as avg_hours
        FROM trainers t
        JOIN working_hours wh ON t.id=wh.trainer_id
        LEFT JOIN assignments a ON wh.assignment_id=a.id
    '''
    conds, params = [], []
    if trainer_id:
        conds.append("t.id=?"); params.append(trainer_id)
    if start_date:
        conds.append("wh.date>=?"); params.append(start_date)
    if end_date:
        conds.append("wh.date<=?"); params.append(end_date)
    if conds:
        query += " WHERE " + " AND ".join(conds)
    query += " GROUP BY t.id, t.name ORDER BY total_hours DESC"

    conn = get_db()
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    conn.close()

    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trainer Report"

        header_fill = PatternFill("solid", fgColor="1a2744")
        header_font = Font(color="39FF14", bold=True, size=11)
        headers = ['Trainer Name', 'Specializations', 'Total Hours', 'Classes Taught', 'Avg Hours/Session']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for ri, row in enumerate(rows, 2):
            ws.cell(ri, 1, row['name'])
            ws.cell(ri, 2, row['specializations'])
            ws.cell(ri, 3, row['total_hours'])
            ws.cell(ri, 4, row['classes_taught'])
            ws.cell(ri, 5, row['avg_hours'])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name='trainer_report.xlsx',
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(['Trainer Name', 'Specializations', 'Total Hours', 'Classes Taught', 'Avg Hours/Session'])
        for row in rows:
            w.writerow([row['name'], row['specializations'], row['total_hours'], row['classes_taught'], row['avg_hours']])
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode()),
                         download_name='trainer_report.csv',
                         as_attachment=True,
                         mimetype='text/csv')

@app.route('/api/dashboard')
def dashboard():
    conn = get_db()
    trainers   = conn.execute("SELECT COUNT(*) as c FROM trainers WHERE status='active'").fetchone()['c']
    classes    = conn.execute("SELECT COUNT(*) as c FROM classes").fetchone()['c']
    assign     = conn.execute("SELECT COUNT(*) as c FROM assignments").fetchone()['c']
    hours_sum  = conn.execute("SELECT ROUND(SUM(hours_worked),1) as s FROM working_hours").fetchone()['s'] or 0
    recent     = conn.execute('''
        SELECT t.name, c.name as cls, a.assigned_date, a.start_time, a.end_time, a.status
        FROM assignments a JOIN trainers t ON a.trainer_id=t.id JOIN classes c ON a.class_id=c.id
        ORDER BY a.assigned_date DESC LIMIT 5
    ''').fetchall()
    conn.close()
    return jsonify({
        'trainers': trainers, 'classes': classes,
        'assignments': assign, 'total_hours': hours_sum,
        'recent': [dict(r) for r in recent]
    })

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
